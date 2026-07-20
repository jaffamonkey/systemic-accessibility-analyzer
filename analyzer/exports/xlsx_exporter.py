import re
from collections import defaultdict
import json
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule

from services.bi_fields import clean_page_name, humanize_slug, humanize_page_key, severity_sort_value, issue_scope_sort_value
from services.severity import normalize_severity
from services.wcag_refs import enrich_wcag_rule, WCAG_SUCCESS_CRITERIA

def _looks_like_wcag_code(value):
    if not value:
        return False

    text = str(value).strip()

    # Plain WCAG code, optionally with technique tag like [G17]
    if re.fullmatch(r"\d+\.\d+\.\d+(?:\s*\[[A-Z0-9]+\])?", text):
        return True

    # HTML_CodeSniffer style rule ids
    if text.startswith("WCAG2"):
        return True

    return False


def _humanize_rule_id(rule_id):
    if not rule_id:
        return ""
    text = str(rule_id).strip().replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text.title()


ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def _excel_safe(value):
    if value is None:
        return ""

    text = str(value)
    text = ILLEGAL_EXCEL_RE.sub("", text)

    # Excel cell limit
    if len(text) > 32767:
        text = text[:32764] + "..."

    return text


def auto_width(ws):
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)


def style_sheet(ws):
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    row_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = row_fill
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_severity_colors(ws, col_letter):
    severity_colors = {
        "critical": "E15759",
        "serious": "F28E2B",
        "moderate": "EDC949",
        "minor": "59A14F",
        "warning": "E9D66B",
        "unknown": "C9CED6",
    }

    for row in ws.iter_rows(min_row=2):
        cell = row[ord(col_letter) - 65]
        value = str(cell.value).lower() if cell.value else ""
        color = severity_colors.get(value)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_wcag_colors(ws, col_letter):
    wcag_colors = {"A": "C6E0B4", "AA": "FFD966", "AAA": "F4B183"}
    for row in ws.iter_rows(min_row=2):
        cell = row[ord(col_letter) - 65]
        value = str(cell.value) if cell.value else ""
        color = wcag_colors.get(value)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _write_sheet(ws, headers, rows, severity_col=None, wcag_level_col=None):
    ws.append([_excel_safe(v) for v in headers])
    for row in rows:
        ws.append([_excel_safe(v) for v in row])
    style_sheet(ws)
    auto_width(ws)
    if severity_col:
        apply_severity_colors(ws, severity_col)
    if wcag_level_col:
        apply_wcag_colors(ws, wcag_level_col)




def _resolve_rule_display(row):
    """
    Prefer descriptive rule labels. Avoid showing a bare WCAG code in the Rule column
    when WCAG already has its own column.
    """
    for key in ("rule_name", "rule_label", "title"):
        value = row.get(key)
        if value and not _looks_like_wcag_code(value):
            return value

    rule_id = row.get("ruleId") or row.get("rule_id")
    if rule_id and not _looks_like_wcag_code(rule_id):
        return _humanize_rule_id(rule_id)

    wcag = row.get("wcag")
    wcag_title = row.get("wcag_title")
    if wcag_title:
        return str(wcag_title)

    if wcag:
        meta = WCAG_SUCCESS_CRITERIA.get(str(wcag).strip(), {})
        title = meta.get("title")
        if title:
            return title
        return str(wcag)

    if rule_id:
        return str(rule_id)

    value = row.get("rule")
    if value:
        return str(value)

    return ""

def _resolve_wcag_name(row):
    wcag = row.get("wcag")
    if not wcag:
        return ""
    meta = WCAG_SUCCESS_CRITERIA.get(str(wcag).strip(), {})
    return meta.get("title", "")


def _build_powerbi_rows(rows, clusters):
    cluster_map = {}
    for c in clusters:
        key = (
            c.get("wcag"),
            c.get("component"),
            normalize_severity(c.get("severity")),
        )
        cluster_map.setdefault(key, c)

    dataset = []
    for idx, r in enumerate(rows, start=1):
        key = (
            r.get("wcag"),
            r.get("component"),
            normalize_severity(r.get("severity")),
        )
        cluster = cluster_map.get(key, {})
        dataset.append({
            "finding_key": idx,
            "page": r.get("page"),
            "page_display": r.get("page_display") or humanize_page_key(r.get("page")) or clean_page_name(r.get("page")),
            "page_group": r.get("page_group") or "root",
            "page_names": cluster.get("page_names") or (r.get("page_display") or humanize_page_key(r.get("page")) or clean_page_name(r.get("page"))),
            "rule": _resolve_rule_display(r),
            "wcag_name": _resolve_wcag_name(r),
            "rule_id": r.get("ruleId"),
            "wcag": r.get("wcag"),
            "wcag_level": r.get("wcag_level"),
            "wcag_level_sort": r.get("wcag_level_sort"),
            "severity": normalize_severity(r.get("severity")),
            "severity_sort": r.get("severity_sort"),
            "severity_label": r.get("severity_label") or humanize_slug(r.get("severity")),
            "component": r.get("component"),
            "component_display": r.get("component_display") or humanize_slug(r.get("component")),
            "component_group": r.get("component_group"),
            "component_group_display": r.get("component_group_display") or humanize_slug(r.get("component_group")),
            "display_pattern": r.get("display_pattern") or humanize_slug(r.get("pattern")),
            "pattern": r.get("pattern"),
            "design_system": r.get("design_system"),
            "design_system_issue": bool(cluster.get("design_system_issue") or r.get("design_system_issue")),
            "issue_scope": cluster.get("issue_scope") or r.get("issue_scope") or "Unknown",
            "issue_scope_sort": cluster.get("issue_scope_sort") or r.get("issue_scope_sort") or issue_scope_sort_value(cluster.get("issue_scope") or r.get("issue_scope")),
            "is_systemic": bool(cluster.get("systemic") or r.get("is_systemic")),
            "affected_pages_count": cluster.get("affected_pages_count") or r.get("affected_pages_count") or 1,
            "issue_rank_score": cluster.get("issue_rank_score") or r.get("issue_rank_score") or 0,
            "owner_team": cluster.get("owner_team") or r.get("owner_team"),
            "root_cause": cluster.get("root_cause"),
            "instance_count": r.get("instance_count"),
            "tool_count": r.get("tool_count"),
            "tool_family_count": r.get("tool_family_count"),
            "tool_families": ", ".join(r.get("tool_families", [])),
            "consensus": r.get("consensus"),
            "confidence": r.get("confidence"),
            "message": r.get("message"),
            "dom": r.get("dom"),
            "dom_path": r.get("dom_path") or r.get("selector") or r.get("dom"),
            "fingerprint": r.get("fingerprint") or cluster.get("fingerprint"),
            "sources": ", ".join(r.get("sources", [])),
            "source": r.get("source"),
        })
    return dataset


def _build_star_schema(powerbi_rows):
    def ordered_unique(rows, key_field):
        seen = set()
        items = []
        for row in rows:
            key = row.get(key_field)
            if key in seen:
                continue
            seen.add(key)
            items.append(row)
        return items

    pattern_stats = {}
    for row in powerbi_rows:
        pattern = row.get("pattern")
        if not pattern:
            continue
        stats = pattern_stats.setdefault(pattern, {
            "display_pattern": row.get("display_pattern"),
            "design_system": row.get("design_system"),
            "design_system_issue": False,
            "issue_scope": row.get("issue_scope") or "Unknown",
            "issue_scope_sort": row.get("issue_scope_sort") or issue_scope_sort_value(row.get("issue_scope")),
            "is_systemic": False,
            "affected_pages_count": 0,
            "issue_rank_score": 0,
            "root_cause": row.get("root_cause"),
            "component": row.get("component"),
            "component_display": row.get("component_display"),
            "owner_team": row.get("owner_team"),
            "severity": row.get("severity"),
            "severity_display": row.get("severity_label") or humanize_slug(row.get("severity")),
            "severity_sort": row.get("severity_sort") or severity_sort_value(row.get("severity")),
            "findings_count": 0,
            "page_keys": set(),
        })
        stats["findings_count"] += 1
        stats["design_system_issue"] = stats["design_system_issue"] or bool(row.get("design_system_issue"))
        stats["is_systemic"] = stats["is_systemic"] or bool(row.get("is_systemic"))
        if (row.get("issue_scope_sort") or 99) < (stats.get("issue_scope_sort") or 99):
            stats["issue_scope"] = row.get("issue_scope") or stats.get("issue_scope")
            stats["issue_scope_sort"] = row.get("issue_scope_sort") or stats.get("issue_scope_sort")
        stats["affected_pages_count"] = max(stats["affected_pages_count"], row.get("affected_pages_count") or 0)
        stats["issue_rank_score"] = max(stats["issue_rank_score"], row.get("issue_rank_score") or 0)
        page = row.get("page")
        if page:
            stats["page_keys"].add(page)
        current_sort = row.get("severity_sort") or severity_sort_value(row.get("severity"))
        if current_sort < stats["severity_sort"]:
            stats["severity_sort"] = current_sort
            stats["severity"] = row.get("severity")
            stats["severity_display"] = row.get("severity_label") or humanize_slug(row.get("severity"))

    ranked_patterns = sorted(
        pattern_stats.items(),
        key=lambda item: (
            item[1]["issue_rank_score"],
            item[1]["affected_pages_count"] or len(item[1]["page_keys"]),
            item[1]["findings_count"],
            -item[1]["severity_sort"],
        ),
        reverse=True,
    )
    for rank, (_, stats) in enumerate(ranked_patterns, start=1):
        stats["top_fix_rank"] = rank
        stats["top_fix_candidate"] = rank <= 10
        if not stats["affected_pages_count"]:
            stats["affected_pages_count"] = len(stats["page_keys"])

    page_dim = []
    page_key_map = {}
    for idx, row in enumerate(ordered_unique(powerbi_rows, "page"), start=1):
        page_key_map[row.get("page")] = idx
        page_dim.append({
            "page_key": idx,
            "page": row.get("page"),
            "page_display": row.get("page_display"),
            "page_group": row.get("page_group"),
        })

    rule_dim = []
    rule_key_map = {}
    for idx, row in enumerate(ordered_unique(powerbi_rows, "rule_id"), start=1):
        rule_code = row.get("rule_id") or row.get("wcag") or row.get("rule")
        rule_key_map[rule_code] = idx
        rule_dim.append({
            "rule_key": idx,
            "rule_id": row.get("rule_id"),
            "rule": row.get("rule"),
            "wcag": row.get("wcag"),
            "wcag_name": row.get("wcag_name"),
            "wcag_level": row.get("wcag_level"),
            "wcag_level_sort": row.get("wcag_level_sort"),
        })

    component_dim = []
    component_key_map = {}
    for idx, row in enumerate(ordered_unique(powerbi_rows, "component"), start=1):
        component_key_map[row.get("component")] = idx
        component_dim.append({
            "component_key": idx,
            "component": row.get("component"),
            "component_display": row.get("component_display"),
            "component_group": row.get("component_group"),
            "component_group_display": row.get("component_group_display"),
            "owner_team": row.get("owner_team"),
        })

    pattern_dim = []
    pattern_key_map = {}
    for idx, row in enumerate(ordered_unique(powerbi_rows, "pattern"), start=1):
        pattern = row.get("pattern")
        pattern_key_map[pattern] = idx
        stats = pattern_stats.get(pattern, {})
        pattern_dim.append({
            "pattern_key": idx,
            "pattern": pattern,
            "display_pattern": row.get("display_pattern"),
            "design_system": row.get("design_system"),
            "design_system_issue": stats.get("design_system_issue", row.get("design_system_issue")),
            "is_systemic": stats.get("is_systemic", row.get("is_systemic")),
            "affected_pages_count": stats.get("affected_pages_count", row.get("affected_pages_count")),
            "issue_rank_score": stats.get("issue_rank_score", row.get("issue_rank_score")),
            "root_cause": row.get("root_cause"),
            "component_display": stats.get("component_display"),
            "owner_team": stats.get("owner_team"),
            "severity": stats.get("severity"),
            "severity_display": stats.get("severity_display"),
            "severity_sort": stats.get("severity_sort"),
            "findings_count": stats.get("findings_count", 0),
            "top_fix_rank": stats.get("top_fix_rank"),
            "top_fix_candidate": stats.get("top_fix_candidate"),
        })

    fact_findings = []
    for row in powerbi_rows:
        rule_code = row.get("rule_id") or row.get("wcag") or row.get("rule")
        stats = pattern_stats.get(row.get("pattern"), {})
        fact_findings.append({
            "finding_key": row.get("finding_key"),
            "page_key": page_key_map.get(row.get("page")),
            "rule_key": rule_key_map.get(rule_code),
            "component_key": component_key_map.get(row.get("component")),
            "pattern_key": pattern_key_map.get(row.get("pattern")),
            "severity": row.get("severity"),
            "severity_sort": row.get("severity_sort"),
            "severity_label": row.get("severity_label"),
            "design_system_issue": row.get("design_system_issue"),
            "is_systemic": row.get("is_systemic"),
            "affected_pages_count": row.get("affected_pages_count"),
            "issue_rank_score": row.get("issue_rank_score"),
            "pattern_findings_count": stats.get("findings_count", 1),
            "top_fix_rank": stats.get("top_fix_rank"),
            "top_fix_candidate": stats.get("top_fix_candidate"),
            "instance_count": row.get("instance_count"),
            "tool_count": row.get("tool_count"),
            "tool_family_count": row.get("tool_family_count"),
            "tool_families": row.get("tool_families"),
            "consensus": row.get("consensus"),
            "confidence": row.get("confidence"),
            "source": row.get("source"),
            "sources": row.get("sources"),
            "message": row.get("message"),
            "dom": row.get("dom"),
            "dom_path": row.get("dom_path"),
            "fingerprint": row.get("fingerprint"),
        })

    fix_panel = []
    for pattern, stats in ranked_patterns[:10]:
        pattern_key = pattern_key_map.get(pattern)
        component_key = component_key_map.get(stats.get("component"))
        fix_panel.append({
            "top_fix_rank": stats.get("top_fix_rank"),
            "pattern_key": pattern_key,
            "component_key": component_key,
            "display_pattern": stats.get("display_pattern"),
            "component_display": stats.get("component_display"),
            "severity_display": stats.get("severity_display"),
            "severity_sort": stats.get("severity_sort"),
            "findings_count": stats.get("findings_count"),
            "affected_pages_count": stats.get("affected_pages_count"),
            "systemic_label": "Yes" if stats.get("is_systemic") else "No",
            "is_systemic": stats.get("is_systemic"),
            "owner_team": stats.get("owner_team"),
            "priority_score": stats.get("issue_rank_score"),
            "root_cause": stats.get("root_cause"),
        })

    return {
        "fact_findings": fact_findings,
        "dim_page": page_dim,
        "dim_rule": rule_dim,
        "dim_component": component_dim,
        "dim_pattern": pattern_dim,
        "fix_panel": fix_panel,
    }




def _add_fix_impact_chart(ws, data_sheet_name):
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Fix-once impact by pattern"
    chart.y_axis.title = "Pattern"
    chart.x_axis.title = "Affected pages"
    chart.height = 8
    chart.width = 18

    data = Reference(ws.parent[data_sheet_name], min_col=9, min_row=1, max_row=min(11, ws.parent[data_sheet_name].max_row))
    cats = Reference(ws.parent[data_sheet_name], min_col=4, min_row=2, max_row=min(11, ws.parent[data_sheet_name].max_row))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    ws.add_chart(chart, "B2")


def _add_owner_team_chart(ws, data_sheet_name):
    source = ws.parent[data_sheet_name]
    owner_counts = defaultdict(int)
    for row in source.iter_rows(min_row=2, values_only=True):
        owner = row[11] or "Unassigned"
        findings = row[7] or 0
        owner_counts[str(owner)] += int(findings)

    owner_sheet = ws.parent.create_sheet("Chart Data - Owners")
    owner_sheet.append(["Owner Team", "Findings"])
    for owner, total in sorted(owner_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
        owner_sheet.append([owner, total])

    chart = BarChart()
    chart.style = 10
    chart.title = "Owner team workload"
    chart.y_axis.title = "Findings"
    chart.x_axis.title = "Owner team"
    chart.height = 7
    chart.width = 16

    data = Reference(owner_sheet, min_col=2, min_row=1, max_row=owner_sheet.max_row)
    cats = Reference(owner_sheet, min_col=1, min_row=2, max_row=owner_sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    ws.add_chart(chart, "B20")


def _add_page_pareto_sheet(wb, powerbi_data):
    counts = defaultdict(int)
    for row in powerbi_data:
        counts[row.get("page_display") or row.get("page") or "Unknown"] += 1

    sorted_pages = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:20]
    ws = wb.create_sheet("Page Pareto")
    ws.append(["Page", "Findings", "Cumulative Findings", "Cumulative %"])
    running = 0
    total = sum(v for _, v in sorted_pages) or 1
    for page, value in sorted_pages:
        running += value
        ws.append([page, value, running, running / total])

    style_sheet(ws)
    auto_width(ws)
    for cell in ws["D"][1:]:
        cell.number_format = "0.0%"

    bar = BarChart()
    bar.title = "Page concentration Pareto"
    bar.y_axis.title = "Findings"
    bar.x_axis.title = "Page"
    bar.height = 8
    bar.width = 18
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = True

    line = LineChart()
    line.y_axis.title = "Cumulative %"
    line.y_axis.axId = 200
    line.height = 8
    line.width = 18
    line_data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.crosses = "max"
    line.y_axis.number_format = "0%"
    line.legend = None

    bar += line
    ws.add_chart(bar, "F2")
    return ws


def _add_component_severity_heatmap(wb, powerbi_data):
    matrix = defaultdict(lambda: defaultdict(int))
    severities = ["critical", "serious", "moderate", "minor", "warning", "unknown"]

    for row in powerbi_data:
        comp = row.get("component_display") or row.get("component") or "Other"
        sev = (row.get("severity") or "unknown").lower()
        if sev not in severities:
            sev = "unknown"
        matrix[comp][sev] += 1

    components = sorted(matrix.keys(), key=lambda c: sum(matrix[c].values()), reverse=True)[:20]
    ws = wb.create_sheet("Component Severity Heatmap")
    ws.append(["Component"] + [s.title() for s in severities])

    for comp in components:
        ws.append([comp] + [matrix[comp][s] for s in severities])

    style_sheet(ws)
    auto_width(ws)

    if ws.max_row >= 2:
        rng = f"B2:G{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFF7FB",
                mid_type="percentile", mid_value=50, mid_color="9ECAE1",
                end_type="max", end_color="3182BD"
            )
        )
    return ws



def _add_component_wcag_matrix(wb, powerbi_data):
    matrix = defaultdict(lambda: defaultdict(int))

    for row in powerbi_data:
        comp = row.get("component_display") or row.get("component") or "Other"
        wcag = row.get("wcag") or "Unmapped"
        matrix[comp][wcag] += 1

    all_wcag = sorted(
        {wcag for comp_map in matrix.values() for wcag in comp_map.keys()},
        key=lambda x: (x == "Unmapped", x)
    )
    top_components = sorted(matrix.keys(), key=lambda c: sum(matrix[c].values()), reverse=True)[:20]
    top_wcag = all_wcag[:15]

    ws = wb.create_sheet("Component WCAG Matrix")
    ws.append(["Component"] + top_wcag)

    for comp in top_components:
        ws.append([comp] + [matrix[comp].get(w, 0) for w in top_wcag])

    style_sheet(ws)
    auto_width(ws)

    if ws.max_row >= 2 and ws.max_column >= 2:
        end_col = get_column_letter(ws.max_column)
        rng = f"B2:{end_col}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="F7FBFF",
                mid_type="percentile", mid_value=50, mid_color="C6DBEF",
                end_type="max", end_color="2171B5"
            )
        )
    return ws

def _add_workbook_charts_sheet(wb, fix_panel_rows, powerbi_data):
    if "Workbook Charts" in wb.sheetnames:
        del wb["Workbook Charts"]
    ws = wb.create_sheet("Workbook Charts")
    ws["A1"] = "Workbook Charts"
    ws["A2"] = "These charts summarize remediation opportunity, ownership, and concentration."
    ws["A1"].font = Font(bold=True, size=16)

    if "Fix Once Benefit Many" in wb.sheetnames and wb["Fix Once Benefit Many"].max_row >= 2:
        _add_fix_impact_chart(ws, "Fix Once Benefit Many")
        _add_owner_team_chart(ws, "Fix Once Benefit Many")

    return ws


def export_xlsx(rows, clusters, metrics, output_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_rows = [
        ["Metric", "Value"],
        ["Violations", metrics.get("violations", 0)],
        ["Pages Affected", metrics.get("pages", 0)],
        # ... your other metrics ...
    ]

    # Move the loop HERE so 'row' is defined
    for row in summary_rows:
        # Sanitize 'row' right here, inside the loop
        sanitized_row = []
        for item in row:
            if item is None:
                sanitized_row.append("")
            elif isinstance(item, list):
                sanitized_row.append(", ".join(str(i) for i in item if i is not None))
            else:
                sanitized_row.append(item)
        
        # Append the now-sanitized row
        ws_summary.append(sanitized_row)

    style_sheet(ws_summary)
    auto_width(ws_summary)

    cluster_headers = [
        "Rule", "WCAG", "Level", "Level Sort", "Severity", "Severity Sort",
        "Component", "Component Group", "Display Pattern", "Systemic",
        "Design System Issue", "Pages", "Affected Pages", "Count", "Issue Rank Score",
        "Owner Team", "Root Cause", "DOM Path", "Fingerprint", "Message", "Sources"
    ]
    
    cluster_rows = []
    for c in clusters:
        # 🔥 THE EXPORT BOUNCER: Drop third-party noise before it becomes a row
        if c.get("component") == "third_party":
            continue

        cluster_rows.append([
            _resolve_rule_display(c),
            c.get("wcag"),
            c.get("wcag_level"),
            c.get("wcag_level_sort"),
            normalize_severity(c.get("severity")), # (Assuming this is what your truncated code said!)
            c.get("severity_sort"),
            c.get("component_display") or humanize_slug(c.get("component")),
            c.get("component_group_display") or humanize_slug(c.get("component_group")),
            c.get("display_pattern"),
            c.get("systemic"),
            c.get("design_system_issue"),
            c.get("pages"),
            c.get("page_names"),
            c.get("count"),
            c.get("issue_rank_score"),
            c.get("owner_team"),
            c.get("root_cause"),
            c.get("dom_path") or c.get("dom"),
            c.get("fingerprint"),
            c.get("message"),
            ", ".join(c.get("sources", [])),
        ])
        
    ws_clusters = wb.create_sheet("Systemic Clusters")
    _write_sheet(ws_clusters, cluster_headers, cluster_rows, severity_col="E", wcag_level_col="C")

    powerbi_data = _build_powerbi_rows(rows, clusters)

    details_headers = [
        "Page", "Page Group", "Affected Pages", "Rule", "Rule Id", "WCAG", "WCAG Name", "Level", "Level Sort",
        "Severity", "Severity Sort", "Component", "Component Group", "Display Pattern",
        "Design System", "Design System Issue", "Systemic", "Affected Pages Count",
        "Issue Rank Score", "Owner Team", "Consensus", "Confidence", "Instances",
        "Tool Count", "DOM", "DOM Path", "Fingerprint", "Message", "Sources"
    ]
    detail_rows = []
    for r in powerbi_data:
        detail_rows.append([
            r["page_display"],
            r["page_group"],
            r.get("page_names"),
            r["rule"],
            r["rule_id"],
            r["wcag"],
            r.get("wcag_name"),
            r["wcag_level"],
            r["wcag_level_sort"],
            r["severity"],
            r["severity_sort"],
            r["component_display"],
            r["component_group_display"],
            r["display_pattern"],
            r["design_system"],
            r["design_system_issue"],
            r["is_systemic"],
            r["affected_pages_count"],
            r["issue_rank_score"],
            r["owner_team"],
            r["consensus"],
            r["confidence"],
            r["instance_count"],
            r["tool_count"],
            r["dom"],
            r.get("dom_path"),
            r.get("fingerprint"),
            r["message"],
            r["sources"],
        ])
    ws_details = wb.create_sheet("Issue Details")
    _write_sheet(ws_details, details_headers, detail_rows, severity_col="J", wcag_level_col="H")

    powerbi_headers = [
        "finding_key", "page", "page_display", "page_group", "rule", "rule_id", "wcag", "wcag_name", "wcag_level",
        "wcag_level_sort", "severity", "severity_sort", "severity_label", "component",
        "component_display", "component_group", "component_group_display", "pattern",
        "display_pattern", "design_system", "design_system_issue", "issue_scope", "issue_scope_sort", "is_systemic",
        "affected_pages_count", "issue_rank_score", "owner_team", "root_cause",
        "instance_count", "tool_count", "consensus", "confidence", "message", "dom",
        "dom_path", "fingerprint", "sources", "source"
    ]
    powerbi_rows = []
    for r in powerbi_data:
        powerbi_rows.append([r.get(h) for h in powerbi_headers])
    ws_powerbi = wb.create_sheet("Power BI Findings")
    _write_sheet(ws_powerbi, powerbi_headers, powerbi_rows, severity_col="J", wcag_level_col="H")

    pattern_rollup = defaultdict(lambda: {
        "findings": 0,
        "affected_pages_count": 0,
        "max_rank": 0,
        "systemic": False,
        "design_system_issue": False,
    })
    for c in clusters:
        if c.get("component") == "third_party":
            continue

        key = (
            c.get("display_pattern"),
            c.get("component_display") or humanize_slug(c.get("component")),
            c.get("wcag"),
            normalize_severity(c.get("severity")),
            c.get("owner_team"),
            c.get("root_cause"),
        )
        pattern_rollup[key]["findings"] += c.get("count", 0)
        pattern_rollup[key]["affected_pages_count"] += c.get("affected_pages_count", c.get("pages", 0))
        pattern_rollup[key]["max_rank"] = max(pattern_rollup[key]["max_rank"], c.get("issue_rank_score", 0))
        pattern_rollup[key]["systemic"] = pattern_rollup[key]["systemic"] or bool(c.get("systemic"))
        pattern_rollup[key]["design_system_issue"] = pattern_rollup[key]["design_system_issue"] or bool(c.get("design_system_issue"))

    pattern_headers = [
        "display_pattern", "component_display", "wcag", "severity", "severity_sort",
        "findings", "affected_pages_count", "is_systemic", "issue_scope", "design_system_issue",
        "issue_rank_score", "owner_team", "root_cause"
    ]
    pattern_rows = []
    for (display_pattern, component_display, wcag, severity, owner_team, root_cause), data in sorted(
        pattern_rollup.items(),
        key=lambda item: (item[1]["max_rank"], item[1]["findings"], item[1]["affected_pages_count"]),
        reverse=True,
    ):
        pattern_rows.append([
            display_pattern,
            component_display,
            wcag,
            severity,
            {"critical": 1, "serious": 2, "moderate": 3, "minor": 4, "warning": 5}.get(str(severity).lower(), 6),
            data["findings"],
            data["affected_pages_count"],
            data["systemic"],
            "Shared" if data["design_system_issue"] else "Local",
            data["design_system_issue"],
            data["max_rank"],
            owner_team,
            root_cause,
        ])
    ws_patterns = wb.create_sheet("Power BI Patterns")
    _write_sheet(ws_patterns, pattern_headers, pattern_rows, severity_col="D")

    schema = _build_star_schema(powerbi_data)

    fact_headers = [
        "finding_key", "page_key", "rule_key", "component_key", "pattern_key",
        "severity", "severity_sort", "severity_label", "design_system_issue", "issue_scope", "issue_scope_sort", "is_systemic",
        "affected_pages_count", "issue_rank_score", "pattern_findings_count", "top_fix_rank",
        "top_fix_candidate", "instance_count", "tool_count", "consensus", "confidence",
        "source", "sources", "message", "dom", "dom_path", "fingerprint"
    ]
    fact_rows = [[row.get(h) for h in fact_headers] for row in schema["fact_findings"]]
    ws_fact = wb.create_sheet("Fact Findings")
    _write_sheet(ws_fact, fact_headers, fact_rows, severity_col="F")

    dim_page_headers = ["page_key", "page", "page_display", "page_group"]
    dim_page_rows = [[row.get(h) for h in dim_page_headers] for row in schema["dim_page"]]
    ws_dim_page = wb.create_sheet("Dim Page")
    _write_sheet(ws_dim_page, dim_page_headers, dim_page_rows)

    dim_rule_headers = ["rule_key", "rule_id", "rule", "wcag", "wcag_name", "wcag_level", "wcag_level_sort"]
    dim_rule_rows = [[row.get(h) for h in dim_rule_headers] for row in schema["dim_rule"]]
    ws_dim_rule = wb.create_sheet("Dim Rule")
    _write_sheet(ws_dim_rule, dim_rule_headers, dim_rule_rows, wcag_level_col="F")

    dim_component_headers = ["component_key", "component", "component_display", "component_group", "component_group_display", "owner_team"]
    dim_component_rows = [[row.get(h) for h in dim_component_headers] for row in schema["dim_component"]]
    ws_dim_component = wb.create_sheet("Dim Component")
    _write_sheet(ws_dim_component, dim_component_headers, dim_component_rows)

    dim_pattern_headers = [
        "pattern_key", "pattern", "display_pattern", "design_system", "design_system_issue",
        "is_systemic", "affected_pages_count", "issue_rank_score", "root_cause",
        "component_display", "owner_team", "severity", "severity_display", "severity_sort",
        "findings_count", "top_fix_rank", "top_fix_candidate"
    ]
    dim_pattern_rows = [[row.get(h) for h in dim_pattern_headers] for row in schema["dim_pattern"]]
    ws_dim_pattern = wb.create_sheet("Dim Pattern")
    _write_sheet(ws_dim_pattern, dim_pattern_headers, dim_pattern_rows)


    fix_panel_headers = [
        "top_fix_rank", "pattern_key", "component_key", "display_pattern", "component_display",
        "severity_display", "severity_sort", "findings_count", "affected_pages_count",
        "systemic_label", "is_systemic", "owner_team", "priority_score", "root_cause"
    ]
    fix_panel_rows = [[row.get(h) for h in fix_panel_headers] for row in schema["fix_panel"]]
    ws_fix_panel = wb.create_sheet("Fix Once Benefit Many")
    _write_sheet(ws_fix_panel, fix_panel_headers, fix_panel_rows, severity_col="F")

    ws_fixes = wb.create_sheet("Top Fixes")
    fix_headers = ["Fix", "Violations Removed", "Pages Affected"]
    fix_rows = [[fix, data["violations"], data["pages"]] for fix, data in metrics.get("top_fixes", {}).items()]
    _write_sheet(ws_fixes, fix_headers, fix_rows)

    glossary_headers = ["Column", "Description"]
    glossary_rows = [
        ["severity_sort", "Numeric sort key for Power BI severity visuals"],
        ["wcag_level_sort", "Numeric sort key for WCAG level visuals"],
        ["display_pattern", "Short human-readable issue grouping label"],
        ["affected_pages_count", "Number of pages impacted by the grouped issue"],
        ["is_systemic", "True when the issue is repeated across multiple pages"],
        ["design_system_issue", "True when the issue appears design-system related"],
        ["issue_rank_score", "Priority score for sorting and triage"],
        ["owner_team", "Suggested owning team for remediation"],
        ["pattern_findings_count", "How many processed findings belong to the pattern"],
        ["top_fix_rank", "Pre-ranked order for the Fix Once, Benefit Many panel"],
        ["top_fix_candidate", "True for patterns that belong in the top-fix shortlist"],
        ["page_group", "Top-level page grouping for filtering"],
        ["Fact Findings", "Star-schema fact table with one row per processed finding"],
        ["Dim Page", "Page dimension keyed by page_key"],
        ["Dim Rule", "Rule dimension keyed by rule_key"],
        ["Dim Component", "Component dimension keyed by component_key"],
        ["Dim Pattern", "Pattern dimension keyed by pattern_key"],
    ]
    ws_glossary = wb.create_sheet("Data Glossary")
    _write_sheet(ws_glossary, glossary_headers, glossary_rows)

    model_headers = ["Section", "Guidance"]
    model_rows = [
        ["Relationship", "Fact Findings.page_key = Dim Page.page_key"],
        ["Relationship", "Fact Findings.rule_key = Dim Rule.rule_key"],
        ["Relationship", "Fact Findings.component_key = Dim Component.component_key"],
        ["Relationship", "Fact Findings.pattern_key = Dim Pattern.pattern_key"],
        ["Panel Title", "Fix Once, Benefit Many"],
        ["Panel Subtitle", "Prioritized issues with the widest impact across pages and components"],
        ["Panel Rows", "Use Dim Pattern.display_pattern, Dim Component.component_display, and Fact Findings.issue_scope"],
        ["Panel Metrics", "Findings Count = COUNTROWS(Fact Findings)"],
        ["Panel Metrics", "Pages Impacted = MAX(Fact Findings[affected_pages_count])"],
        ["Panel Metrics", "Priority Score = SUM(Fact Findings[issue_rank_score])"],
        ["Panel Metrics", "Systemic Label = IF(MAX(Fact Findings[is_systemic]), 'Yes', 'No')"],
        ["Panel Sorting", "Sort by Priority Score desc, then Pages Impacted desc"],
        ["Panel Filter", "Top N = 10 by Priority Score; exclude blank display_pattern"],
        ["Supporting Card", "Systemic fixes = DISTINCTCOUNT(Fact Findings[pattern_key]) where is_systemic = TRUE"],
        ["Supporting Card", "Pages impacted by top 5 fixes = DISTINCTCOUNT(Fact Findings[page_key]) filtered to top_fix_rank <= 5"],
        ["Supporting Card", "Top owner team = team with highest COUNTROWS(Fact Findings) in current filter context"],
    ]
    ws_model = wb.create_sheet("Model Notes")
    _write_sheet(ws_model, model_headers, model_rows)

    _add_component_severity_heatmap(wb, powerbi_data)
    _add_component_wcag_matrix(wb, powerbi_data)

    wb.save(output_path)
